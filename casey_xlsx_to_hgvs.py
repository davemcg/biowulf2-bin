#!/usr/bin python3

from openpyxl import load_workbook
import argparse
import hgvs.validator
import hgvs.exceptions
import hgvs.dataproviders.uta
import hgvs.parser
import hgvs.assemblymapper

parser = argparse.ArgumentParser(description= 'Takes in xlsx files from John Chiang / MVLGenomics Gene Panels and does the following:\n 1. Identifies hidden rows and skips these \n 2. Extracts hgvs \n\n Usage: casey_xlsx_to_hgvs.py YOURFILE.xlsx > output.tsv')

parser.add_argument('xlsx_file', help = 'John Chiang / OSU / MVL patient report xlsx file')

args = parser.parse_args()

wb = load_workbook(args.xlsx_file)
ws = wb[wb.get_sheet_names()[0]]
# identify header row, HGVS coding column, and zygosity column
for row in range(2, ws.max_row):
    for column in 'ABCDEFG':
        cell_name = "{}{}".format(column, row)
        if ws[cell_name].value == 'HGVScoding':
            header_row = row
            HGVS_column = column
        if ws[cell_name].value == 'Zygosity':
            Zygosity_column = column

# get HGVS and zygosity, skipping hidden rows
for row in range(header_row+1, ws.max_row):
    if ws.row_dimensions[row].visible is True:
        cell_name_hgvs = "{}{}".format(HGVS_column, row)
        cell_name_zyg = "{}{}".format(Zygosity_column, row)
        if ws[cell_name_hgvs].value is not None and ws[cell_name_hgvs].value[0:2] == 'NM':
            print(ws[cell_name_hgvs].value, '\t', ws[cell_name_zyg].value)

